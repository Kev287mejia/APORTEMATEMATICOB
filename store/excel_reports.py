import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count, Avg

from .models import DigitalPurchase, ActivationCode, PhysicalOrder, ContactMessage


def get_excel_styles():
    """Retorna estilos consistentes con la paleta de la plataforma."""
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    kpi_title_fill = PatternFill(start_color="FAAB9F", end_color="FAAB9F", fill_type="solid")
    kpi_title_font = Font(name="Calibri", size=11, bold=True, color="2C2C2C")
    kpi_val_font = Font(name="Calibri", size=13, bold=True, color="1A2744")
    total_fill = PatternFill(start_color="F5F0EB", end_color="F5F0EB", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True, color="2C2C2C")
    data_font = Font(name="Calibri", size=10)
    
    border_thin = Border(
        left=Side(style='thin', color='E0D8CF'),
        right=Side(style='thin', color='E0D8CF'),
        top=Side(style='thin', color='E0D8CF'),
        bottom=Side(style='thin', color='E0D8CF')
    )
    border_total = Border(
        top=Side(style='thin', color='1A2744'),
        bottom=Side(style='double', color='1A2744')
    )

    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'kpi_title_fill': kpi_title_fill,
        'kpi_title_font': kpi_title_font,
        'kpi_val_font': kpi_val_font,
        'total_fill': total_fill,
        'total_font': total_font,
        'data_font': data_font,
        'border_thin': border_thin,
        'border_total': border_total
    }


def autofit_columns(ws, start_row=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= start_row and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_master_executive_workbook():
    """Genera el Libro Maestro de Excel con KPIs ejecutivos y desglose por pestañas."""
    styles = get_excel_styles()
    wb = Workbook()

    # ==========================================
    # PESTAÑA 1: RESUMEN EJECUTIVO Y KPIS
    # ==========================================
    ws_kpi = wb.active
    ws_kpi.title = "Resumen Ejecutivo y KPIs"
    ws_kpi.views.sheetView[0].showGridLines = True

    # Cálculos de Métricas
    total_digital_count = DigitalPurchase.objects.count()
    completed_purchases = DigitalPurchase.objects.filter(status='completed')
    total_revenue_usd = completed_purchases.aggregate(Sum('amount'))['amount__sum'] or 0.0
    avg_ticket = completed_purchases.aggregate(Avg('amount'))['amount__avg'] or 0.0
    
    total_orders_count = PhysicalOrder.objects.count()
    total_physical_books = PhysicalOrder.objects.aggregate(Sum('quantity'))['quantity__sum'] or 0

    total_codes = ActivationCode.objects.count()
    used_codes = ActivationCode.objects.filter(is_used=True).count()
    pending_codes = total_codes - used_codes
    activation_rate = (used_codes / total_codes * 100) if total_codes > 0 else 0.0

    total_messages = ContactMessage.objects.count()

    # Título Principal
    ws_kpi.merge_cells("A1:F1")
    t_cell = ws_kpi.cell(row=1, column=1, value="INFORME EJECUTIVO Y ESTADÍSTICAS GENERALES")
    t_cell.font = Font(name="Calibri", size=16, bold=True, color="1A2744")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[1].height = 35

    ws_kpi.merge_cells("A2:F2")
    sub_cell = ws_kpi.cell(row=2, column=1, value=f"Obra: Un Aporte Matemático en el Siglo 21 — Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[2].height = 20

    # Sección 1: Tarjetas de KPIs Principales
    kpi_data = [
        ("TOTAL INGRESOS DIGITALES", f"${float(total_revenue_usd):,.2f} USD", "Ventas completadas en PayPal y Kash"),
        ("VENTAS DIGITALES REGISTRADAS", f"{total_digital_count}", "Transacciones totales procesadas"),
        ("TICKET PROMEDIO POR COMPRA", f"${float(avg_ticket):,.2f} USD", "Promedio de ingreso por cliente"),
        ("LIBROS FÍSICOS SOLICITADOS", f"{total_physical_books} unidades", f"Distribuidos en {total_orders_count} pedidos"),
        ("CÓDIGOS DE ACTIVACIÓN EMITIDOS", f"{total_codes}", f"{used_codes} activados | {pending_codes} disponibles"),
        ("TASA DE ACTIVACIÓN DIGITAL", f"{activation_rate:.1f}%", "Porcentaje de compradores con acceso activo"),
        ("MENSAJES DE CONTACTO RECIBIDOS", f"{total_messages}", "Consultas web de lectores y docentes"),
    ]

    # Encabezado de tabla KPI
    ws_kpi.cell(row=4, column=1, value="Indicador Clave de Rendimiento (KPI)").fill = styles['header_fill']
    ws_kpi.cell(row=4, column=1).font = styles['header_font']
    ws_kpi.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    ws_kpi.cell(row=4, column=2, value="Valor Actual").fill = styles['header_fill']
    ws_kpi.cell(row=4, column=2).font = styles['header_font']
    ws_kpi.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws_kpi.merge_cells("C4:F4")
    ws_kpi.cell(row=4, column=3, value="Detalle / Observación").fill = styles['header_fill']
    ws_kpi.cell(row=4, column=3).font = styles['header_font']
    ws_kpi.cell(row=4, column=3).alignment = Alignment(horizontal="left", vertical="center")
    ws_kpi.row_dimensions[4].height = 25

    for idx, (title, val, desc) in enumerate(kpi_data, 5):
        c1 = ws_kpi.cell(row=idx, column=1, value=title)
        c1.font = Font(name="Calibri", size=10, bold=True, color="2C2C2C")
        c1.border = styles['border_thin']

        c2 = ws_kpi.cell(row=idx, column=2, value=val)
        c2.font = styles['kpi_val_font']
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = styles['border_thin']

        ws_kpi.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=6)
        c3 = ws_kpi.cell(row=idx, column=3, value=desc)
        c3.font = Font(name="Calibri", size=10, italic=True, color="666666")
        c3.border = styles['border_thin']
        
        ws_kpi.row_dimensions[idx].height = 24

    # Desglose de Ventas por Método de Pago
    start_breakdown = len(kpi_data) + 7
    ws_kpi.merge_cells(start_row=start_breakdown, start_column=1, end_row=start_breakdown, end_column=4)
    sec_title = ws_kpi.cell(row=start_breakdown, column=1, value="DESGLOSE DE VENTAS POR PASARELA")
    sec_title.font = Font(name="Calibri", size=12, bold=True, color="1A2744")
    ws_kpi.row_dimensions[start_breakdown].height = 25

    headers_gw = ["Pasarela", "Transacciones", "Monto Total (USD)", "Estado"]
    for col_idx, h in enumerate(headers_gw, 1):
        c = ws_kpi.cell(row=start_breakdown+1, column=col_idx, value=h)
        c.fill = styles['header_fill']
        c.font = styles['header_font']
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles['border_thin']
    ws_kpi.row_dimensions[start_breakdown+1].height = 22

    gw_summary = DigitalPurchase.objects.values('payment_method').annotate(
        total_count=Count('id'),
        total_sum=Sum('amount')
    )
    for row_offset, gw in enumerate(gw_summary, start_breakdown+2):
        name_gw = "PayPal" if gw['payment_method'] == 'paypal' else "Kash Nicaragua"
        c1 = ws_kpi.cell(row=row_offset, column=1, value=name_gw)
        c1.font = styles['data_font']
        c1.border = styles['border_thin']

        c2 = ws_kpi.cell(row=row_offset, column=2, value=gw['total_count'])
        c2.font = styles['data_font']
        c2.alignment = Alignment(horizontal="center")
        c2.border = styles['border_thin']

        c3 = ws_kpi.cell(row=row_offset, column=3, value=float(gw['total_sum'] or 0.0))
        c3.font = styles['data_font']
        c3.number_format = "$#,##0.00"
        c3.alignment = Alignment(horizontal="right")
        c3.border = styles['border_thin']

        c4 = ws_kpi.cell(row=row_offset, column=4, value="Activo")
        c4.font = styles['data_font']
        c4.alignment = Alignment(horizontal="center")
        c4.border = styles['border_thin']
        ws_kpi.row_dimensions[row_offset].height = 20

    autofit_columns(ws_kpi, start_row=4)

    # ==========================================
    # PESTAÑA 2: VENTAS DIGITALES
    # ==========================================
    ws_sales = wb.create_sheet(title="Ventas Digitales")
    ws_sales.views.sheetView[0].showGridLines = True
    
    ws_sales.merge_cells("A1:I1")
    t_sales = ws_sales.cell(row=1, column=1, value="REGISTRO DETALLADO DE VENTAS DIGITALES")
    t_sales.font = Font(name="Calibri", size=14, bold=True, color="1A2744")
    t_sales.alignment = Alignment(horizontal="center", vertical="center")
    ws_sales.row_dimensions[1].height = 30

    headers_sales = ["ID Transacción", "Correo del Comprador", "Monto (USD)", "Moneda", "Pasarela", "Estado", "¿Activado?", "Código de Activación", "Fecha de Compra"]
    for col_idx, h in enumerate(headers_sales, 1):
        c = ws_sales.cell(row=3, column=col_idx, value=h)
        c.fill = styles['header_fill']
        c.font = styles['header_font']
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles['border_thin']
    ws_sales.row_dimensions[3].height = 24

    purchases = DigitalPurchase.objects.all().order_by('-created_at')
    current_row = 4
    for p in purchases:
        code_str = p.activation_code.code if hasattr(p, 'activation_code') and p.activation_code else "N/A"
        date_str = p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        
        row_vals = [
            p.transaction_id or "Sin ID",
            p.buyer_email or "N/A",
            float(p.amount),
            p.currency,
            p.payment_method.upper(),
            p.get_status_display(),
            "Sí" if p.is_used else "No",
            code_str,
            date_str
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_sales.cell(row=current_row, column=col_idx, value=val)
            cell.font = styles['data_font']
            cell.border = styles['border_thin']
            if col_idx == 3:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
        ws_sales.row_dimensions[current_row].height = 20
        current_row += 1

    # Fila de Totales con fórmula
    if purchases.exists():
        ws_sales.cell(row=current_row, column=1, value="TOTAL VENTAS (USD):").font = styles['total_font']
        ws_sales.cell(row=current_row, column=1).alignment = Alignment(horizontal="right")
        ws_sales.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        
        tot_cell = ws_sales.cell(row=current_row, column=3, value=f"=SUM(C4:C{current_row-1})")
        tot_cell.font = styles['total_font']
        tot_cell.number_format = "$#,##0.00"
        tot_cell.border = styles['border_total']
        ws_sales.row_dimensions[current_row].height = 24

    autofit_columns(ws_sales, start_row=3)

    # ==========================================
    # PESTAÑA 3: PEDIDOS FÍSICOS
    # ==========================================
    ws_orders = wb.create_sheet(title="Pedidos Físicos")
    ws_orders.views.sheetView[0].showGridLines = True

    ws_orders.merge_cells("A1:G1")
    t_ord = ws_orders.cell(row=1, column=1, value="REGISTRO DE PEDIDOS DE LIBROS FÍSICOS")
    t_ord.font = Font(name="Calibri", size=14, bold=True, color="1A2744")
    t_ord.alignment = Alignment(horizontal="center", vertical="center")
    ws_orders.row_dimensions[1].height = 30

    headers_orders = ["ID", "Nombre Completo", "Teléfono / WhatsApp", "Departamento", "Cantidad Libros", "Notas del Pedido", "Fecha del Pedido"]
    for col_idx, h in enumerate(headers_orders, 1):
        c = ws_orders.cell(row=3, column=col_idx, value=h)
        c.fill = styles['header_fill']
        c.font = styles['header_font']
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles['border_thin']
    ws_orders.row_dimensions[3].height = 24

    orders = PhysicalOrder.objects.all().order_by('-created_at')
    current_row = 4
    for o in orders:
        date_str = o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""
        row_vals = [
            o.id,
            o.name,
            o.phone,
            o.department,
            o.quantity,
            o.notes or "Sin notas",
            date_str
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_orders.cell(row=current_row, column=col_idx, value=val)
            cell.font = styles['data_font']
            cell.border = styles['border_thin']
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="right")
        ws_orders.row_dimensions[current_row].height = 20
        current_row += 1

    if orders.exists():
        ws_orders.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws_orders.cell(row=current_row, column=1, value="TOTAL LIBROS PEDIDOS:").font = styles['total_font']
        ws_orders.cell(row=current_row, column=1).alignment = Alignment(horizontal="right")
        
        tot_qty = ws_orders.cell(row=current_row, column=5, value=f"=SUM(E4:E{current_row-1})")
        tot_qty.font = styles['total_font']
        tot_qty.border = styles['border_total']
        ws_orders.row_dimensions[current_row].height = 24

    autofit_columns(ws_orders, start_row=3)

    # ==========================================
    # PESTAÑA 4: CÓDIGOS DE ACTIVACIÓN
    # ==========================================
    ws_codes = wb.create_sheet(title="Códigos de Activación")
    ws_codes.views.sheetView[0].showGridLines = True

    ws_codes.merge_cells("A1:F1")
    t_cd = ws_codes.cell(row=1, column=1, value="LISTADO DE CÓDIGOS DE ACCESO")
    t_cd.font = Font(name="Calibri", size=14, bold=True, color="1A2744")
    t_cd.alignment = Alignment(horizontal="center", vertical="center")
    ws_codes.row_dimensions[1].height = 30

    headers_codes = ["Código de Activación", "Estado", "ID Transacción Vinculada", "Correo del Comprador", "Fecha de Creación", "Fecha de Activación"]
    for col_idx, h in enumerate(headers_codes, 1):
        c = ws_codes.cell(row=3, column=col_idx, value=h)
        c.fill = styles['header_fill']
        c.font = styles['header_font']
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles['border_thin']
    ws_codes.row_dimensions[3].height = 24

    codes = ActivationCode.objects.all().order_by('-created_at')
    for row_idx, cd in enumerate(codes, 4):
        tx_id = cd.purchase.transaction_id if cd.purchase else "Manual"
        buyer = cd.purchase.buyer_email if cd.purchase else "Manual"
        c_date = cd.created_at.strftime("%Y-%m-%d %H:%M") if cd.created_at else ""
        u_date = cd.used_at.strftime("%Y-%m-%d %H:%M") if cd.used_at else "Pendiente"

        row_vals = [
            cd.code,
            "Usado" if cd.is_used else "Disponible",
            tx_id,
            buyer,
            c_date,
            u_date
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_codes.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles['data_font']
            cell.border = styles['border_thin']
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="center")
        ws_codes.row_dimensions[row_idx].height = 20

    autofit_columns(ws_codes, start_row=3)

    # ==========================================
    # PESTAÑA 5: MENSAJES DE CONTACTO
    # ==========================================
    ws_msg = wb.create_sheet(title="Mensajes de Contacto")
    ws_msg.views.sheetView[0].showGridLines = True

    ws_msg.merge_cells("A1:F1")
    t_msg = ws_msg.cell(row=1, column=1, value="REGISTRO DE MENSAJES WEB Y CONSULTAS")
    t_msg.font = Font(name="Calibri", size=14, bold=True, color="1A2744")
    t_msg.alignment = Alignment(horizontal="center", vertical="center")
    ws_msg.row_dimensions[1].height = 30

    headers_msg = ["ID", "Nombre del Remitente", "Correo Electrónico", "Asunto", "Mensaje", "Fecha de Envío"]
    for col_idx, h in enumerate(headers_msg, 1):
        c = ws_msg.cell(row=3, column=col_idx, value=h)
        c.fill = styles['header_fill']
        c.font = styles['header_font']
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles['border_thin']
    ws_msg.row_dimensions[3].height = 24

    messages = ContactMessage.objects.all().order_by('-created_at')
    for row_idx, m in enumerate(messages, 4):
        date_str = m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else ""
        row_vals = [
            m.id,
            m.name,
            m.email,
            m.subject or "Sin Asunto",
            m.message,
            date_str
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_msg.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles['data_font']
            cell.border = styles['border_thin']
        ws_msg.row_dimensions[row_idx].height = 20

    autofit_columns(ws_msg, start_row=3)

    return wb
