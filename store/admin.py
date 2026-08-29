import datetime
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import DigitalPurchase, ActivationCode, PhysicalOrder, ContactMessage

admin.site.site_header = "Aporte Matemático — Panel de Administración"
admin.site.site_title = "Aporte Matemático Admin"
admin.site.index_title = "Gestión de Ventas, Códigos y Mensajes"


def apply_excel_styling(ws, title, headers, rows):
    """Genera una hoja de Excel profesional con diseño Odrin (Azul Marino y Dorado)."""
    # Paleta de colores
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    border_thin = Border(
        left=Side(style='thin', color='E0D8CF'),
        right=Side(style='thin', color='E0D8CF'),
        top=Side(style='thin', color='E0D8CF'),
        bottom=Side(style='thin', color='E0D8CF')
    )

    # Título principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{title} — Aporte Matemático")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1A2744")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo con fecha
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Generado el: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.font = Font(name="Calibri", size=9, italic=True, color="777777")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[4].height = 24

    # Datos
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border_thin
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 20

    # Auto-ajuste de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 4 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


@admin.action(description="📊 Exportar Compras seleccionadas a Excel (.xlsx)")
def export_purchases_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas Digitales"

    headers = ["ID Transacción", "Correo del Comprador", "Monto (USD)", "Moneda", "Pasarela", "Estado", "¿Código Usado?", "Código de Activación", "Fecha de Compra"]
    rows = []
    for p in queryset:
        code_str = p.activation_code.code if hasattr(p, 'activation_code') and p.activation_code else "N/A"
        date_str = p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        rows.append([
            p.transaction_id or "Sin ID",
            p.buyer_email or "N/A",
            float(p.amount),
            p.currency,
            p.payment_method.upper(),
            p.get_status_display(),
            "Sí" if p.is_used else "No",
            code_str,
            date_str
        ])

    apply_excel_styling(ws, "Reporte Oficial de Ventas Digitales", headers, rows)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Reporte_Ventas_PayPal_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin.action(description="📦 Exportar Pedidos Físicos a Excel (.xlsx)")
def export_orders_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos Físicos"

    headers = ["ID", "Nombre Completo", "Teléfono / WhatsApp", "Departamento", "Cantidad de Libros", "Notas del Pedido", "Fecha de Solicitud"]
    rows = []
    for o in queryset:
        date_str = o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""
        rows.append([
            o.id,
            o.name,
            o.phone,
            o.department,
            o.quantity,
            o.notes or "Sin notas",
            date_str
        ])

    apply_excel_styling(ws, "Reporte de Pedidos de Libros Físicos", headers, rows)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Reporte_Pedidos_Fisicos_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin.action(description="🔑 Exportar Códigos a Excel (.xlsx)")
def export_codes_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos de Activación"

    headers = ["Código", "Estado", "ID Transacción Vinculada", "Correo Comprador", "Fecha de Creación", "Fecha de Uso"]
    rows = []
    for c in queryset:
        tx_id = c.purchase.transaction_id if c.purchase else "Generado Manualmente"
        buyer = c.purchase.buyer_email if c.purchase else "Manual"
        created_str = c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""
        used_str = c.used_at.strftime("%Y-%m-%d %H:%M") if c.used_at else "Pendiente"
        rows.append([
            c.code,
            "Usado" if c.is_used else "Disponible",
            tx_id,
            buyer,
            created_str,
            used_str
        ])

    apply_excel_styling(ws, "Listado de Códigos de Activación", headers, rows)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Reporte_Codigos_Activacion_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin.action(description="✉️ Exportar Mensajes a Excel (.xlsx)")
def export_messages_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensajes de Contacto"

    headers = ["ID", "Nombre del Remitente", "Correo Electrónico", "Asunto", "Mensaje", "Fecha de Envío"]
    rows = []
    for m in queryset:
        date_str = m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else ""
        rows.append([
            m.id,
            m.name,
            m.email,
            m.subject or "Sin Asunto",
            m.message,
            date_str
        ])

    apply_excel_styling(ws, "Reporte de Mensajes de Contacto Web", headers, rows)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Reporte_Mensajes_Contacto_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin.register(DigitalPurchase)
class DigitalPurchaseAdmin(admin.ModelAdmin):
    list_display = ('transaction_id', 'buyer_email', 'amount', 'currency', 'status', 'is_used', 'created_at')
    search_fields = ('transaction_id', 'buyer_email')
    list_filter = ('status', 'payment_method', 'is_used', 'created_at')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'
    actions = [export_purchases_to_excel]


@admin.register(ActivationCode)
class ActivationCodeAdmin(admin.ModelAdmin):
    list_display = ('code', 'is_used', 'purchase', 'created_at', 'used_at')
    search_fields = ('code', 'purchase__transaction_id', 'purchase__buyer_email')
    list_filter = ('is_used', 'created_at')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'
    actions = [export_codes_to_excel]


@admin.register(PhysicalOrder)
class PhysicalOrderAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'department', 'quantity', 'created_at')
    search_fields = ('name', 'phone', 'department', 'notes')
    list_filter = ('department', 'created_at')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'
    actions = [export_orders_to_excel]


@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'subject', 'created_at')
    search_fields = ('name', 'email', 'subject', 'message')
    list_filter = ('created_at',)
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'
    actions = [export_messages_to_excel]



